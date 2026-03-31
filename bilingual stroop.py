from psychopy import visual, core, event
import os
import datetime
from openpyxl import Workbook, load_workbook
import random

# Set up paths and filenames
excel_path = r"C:\Users\Ankushi\Documents\Colors stroop\stroop.xlsx"
hindi_images_path = r"C:\Users\Ankushi\Documents\Colors stroop\Colors"

# Define English words and colors (excluding white)
english_words = ["red", "blue", "green", "yellow", "purple", "orange", "pink", "brown", "black"]
colors = ["red", "blue", "green", "yellow", "purple", "orange", "pink", "brown", "black"]

# Set up the window with pixel units
win = visual.Window([800, 600], color="white", units="pix")

# Function to create or open an Excel file
def open_or_create_excel(file_path):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "bilingual Stroop Data"
        headers = ["User Serial No.", "Time of Experiment"] + [f"Reaction Time {i+1}" for i in range(20)]
        ws.append(headers)
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)
    return wb

# Function to get the next user serial number
def get_next_serial_number(file_path):
    if not os.path.exists(file_path):
        return 1
    wb = load_workbook(file_path)
    ws = wb["Stroop Data"]
    last_row = ws.max_row
    if last_row > 1:
        last_serial = ws.cell(row=last_row, column=1).value
        return last_serial + 1 if last_serial else 1
    return 1

# Function to log data into the Excel file
def log_data_to_excel(data, file_path):
    wb = open_or_create_excel(file_path)
    ws = wb["Stroop Data"]
    ws.append(data)
    wb.save(file_path)

# Function to run English text trials
def run_english_trials():
    reaction_times = []
    for word in english_words[:10]:  # Use the first 10 English words
        color = random.choice(colors)
        stim = visual.TextStim(win, text=word, color=color, font="Arial Unicode MS", units="pix", height=50)  # Adjust height in pixels
        stim.draw()
        win.flip()

        # Record reaction time and response
        start_time = core.getTime()
        keys = event.waitKeys(keyList=[color[0]], timeStamped=True)
        reaction_time = core.getTime() - start_time
        if keys:  # Ensure the keypress is captured
            reaction_times.append(reaction_time)
    return reaction_times

# Function to run Hindi image trials with random images
def run_hindi_trials():
    # List all images, excluding those with filenames starting with 's'
    all_images = [f for f in os.listdir(hindi_images_path) if f.endswith(('.png', '.jpg', '.jpeg')) and not f.lower().startswith('s')]
    random_images = random.sample(all_images, 10)  # Select 10 random images

    reaction_times = []
    for image_file in random_images:
        image_path = os.path.join(hindi_images_path, image_file)
        expected_key = image_file[0].lower()  # First letter of the filename as expected key

        # Display image
        stim = visual.ImageStim(win, image=image_path)
        stim.draw()
        win.flip()

        # Record reaction time and response
        start_time = core.getTime()
        keys = event.waitKeys(keyList=[expected_key], timeStamped=True)
        reaction_time = core.getTime() - start_time
        if keys:  # Ensure the keypress is captured
            reaction_times.append(reaction_time)
    return reaction_times

# Main program
def main():
    user_serial_no = get_next_serial_number(excel_path)
    experiment_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Instructions for participants
    instructions = visual.TextStim(win, text="Press the corresponding key to match the ink color of the word.\nIgnore the word meaning. Press any key to start.", color="black", font="Arial Unicode MS")
    instructions.draw()
    win.flip()
    event.waitKeys()

    # Run English trials and collect reaction times
    english_reaction_times = run_english_trials()

    # Optional break between trials
    break_text = visual.TextStim(win, text="Take a short break. Press any key to continue with the next set.", color="black", font="Arial Unicode MS")
    break_text.draw()
    win.flip()
    core.wait(10)
    event.waitKeys()

    # Run Hindi trials and collect reaction times
    hindi_reaction_times = run_hindi_trials()

    # Combine data and prepare for logging
    all_reaction_times = english_reaction_times + hindi_reaction_times
    data_row = [user_serial_no, experiment_time] + all_reaction_times

    # Log data into Excel file
    log_data_to_excel(data_row, excel_path)

    # End of experiment
    end_text = visual.TextStim(win, text="Experiment complete. Thank you!", color="black", font="Arial Unicode MS")
    end_text.draw()
    win.flip()
    core.wait(3)
    win.close()
    core.quit()

if __name__ == "__main__":
    main()
